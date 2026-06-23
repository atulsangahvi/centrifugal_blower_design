"""
Centrifugal Blower Design & Manufacturing Toolkit v12 - SI Units
Static pressure input only. Auto-optimised geometry for backward curved, forward curved and radial blade blowers.

Preliminary engineering tool only. Validate final design by prototype testing, AMCA/ISO test procedure,
dynamic balancing, vibration measurement, CFD/FEA and qualified engineering review before manufacture.
"""
from __future__ import annotations
import io, math, zipfile
from dataclasses import dataclass, asdict, replace
from typing import List, Tuple, Dict

import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt

try:
    import ezdxf
    HAS_EZDXF=True
except Exception:
    HAS_EZDXF=False
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    HAS_REPORTLAB=True
except Exception:
    HAS_REPORTLAB=False

STANDARD_MOTORS_KW=[0.18,0.25,0.37,0.55,0.75,1.1,1.5,2.2,3,4,5.5,7.5,11,15,18.5,22,30,37,45,55,75,90,110,132,160,200,250,315]

BLADE_DEFAULTS={
    "Backward Curved / Backward Inclined":{"beta1":24,"beta2":34,"eta":0.74,"phi":0.18,"psi":0.58,"z":12,"z_range":range(8,17),"beta2_range":[28,32,36,40,44]},
    "Forward Curved":{"beta1":28,"beta2":125,"eta":0.58,"phi":0.16,"psi":0.72,"z":36,"z_range":range(28,49,4),"beta2_range":[115,125,135,145]},
    "Radial Blade":{"beta1":25,"beta2":90,"eta":0.62,"phi":0.12,"psi":0.62,"z":8,"z_range":range(6,13),"beta2_range":[88,90,92]},
}
MATERIALS={
    "Mild Steel IS2062 / S275":{"density":7850,"max_tip":90},
    "Galvanized Steel":{"density":7850,"max_tip":75},
    "Stainless Steel 304":{"density":8000,"max_tip":90},
    "Aluminium 6061-T6":{"density":2700,"max_tip":70},
}

@dataclass
class DutyInput:
    airflow_m3h: float
    static_pressure_pa: float
    air_temp_c: float
    altitude_m: float
    density_kgm3: float
    rpm: float
    blade_type: str
    drive_type: str
    motor_eff: float
    drive_eff: float
    design_margin: float
    beta2_deg: float
    beta1_deg: float
    blade_count: int
    outlet_width_ratio: float
    inlet_diameter_ratio: float
    target_flange_velocity_ms: float
    material: str
    blade_thickness_mm: float
    casing_thickness_mm: float
    shaft_allow_shear_mpa: float

@dataclass
class DesignResult:
    q_m3s: float
    density_kgm3: float
    static_pressure_pa: float
    velocity_pressure_pa: float
    total_pressure_pa: float
    rpm: float
    tip_speed_ms: float
    impeller_od_mm: float
    impeller_id_mm: float
    outlet_width_mm: float
    inlet_width_mm: float
    beta1_deg: float
    beta2_deg: float
    blade_count: int
    slip_factor: float
    flow_coeff_phi: float
    pressure_coeff_psi: float
    estimated_total_eff: float
    air_power_kw: float
    shaft_power_kw: float
    selected_motor_kw: float
    specific_speed_metric: float
    specific_diameter_metric: float
    impeller_meridional_velocity_ms: float
    flange_outlet_velocity_ms: float
    volute_outlet_width_mm: float
    volute_outlet_height_mm: float
    volute_cutoff_clearance_mm: float
    shaft_diameter_mm: float
    approx_impeller_mass_kg: float
    sound_db_a_1m: float
    vibration_risk: str
    vibration_score: float
    design_score: float
    warnings: List[str]

# ---------- engineering calculations ----------
def standard_air_density(temp_c:float, altitude_m:float)->float:
    t0=288.15; p0=101325.0; lapse=0.0065; r=287.05; g=9.80665
    t_alt=max(180.0,t0-lapse*altitude_m)
    p=p0*(t_alt/t0)**(g/(r*lapse))
    return p/(r*(temp_c+273.15))

def selected_motor(kw:float)->float:
    for m in STANDARD_MOTORS_KW:
        if m>=kw: return m
    return STANDARD_MOTORS_KW[-1]

def stodola_slip_factor(z:int,beta2_deg:float,d1d2:float)->float:
    beta=math.radians(beta2_deg)
    sig=1.0-(math.pi*abs(math.sin(beta)))/max(z,3)
    sig*=max(0.88,min(1.02,1-0.08*(d1d2-0.55)))
    return max(0.55,min(0.95,sig))

def estimate_sound_db(inp:DutyInput,res_without_sound=None, shaft_kw=None, tip=None, tp=None, vout=None, pc=None)->float:
    if res_without_sound is not None:
        shaft_kw=res_without_sound.shaft_power_kw; tip=res_without_sound.tip_speed_ms; tp=res_without_sound.total_pressure_pa; vout=res_without_sound.flange_outlet_velocity_ms; pc=blade_pitch_chord_ratio(res_without_sound)
    blade_corr={"Backward Curved / Backward Inclined":-3,"Forward Curved":2,"Radial Blade":5}.get(inp.blade_type,0)
    pitch_corr=4 if (pc or 1)<0.55 else 2 if (pc or 1)<0.70 else 0
    return 62+10*math.log10(max(shaft_kw or 0.1,0.1))+25*math.log10(max(tip or 1,1)/35)+6*math.log10(max(tp or 100,100)/500)+max(0,(vout or 0)-12)*0.8+blade_corr+pitch_corr

def design_blower(inp:DutyInput)->DesignResult:
    q=inp.airflow_m3h/3600
    rho=inp.density_kgm3 if inp.density_kgm3>0 else standard_air_density(inp.air_temp_c, inp.altitude_m)
    defaults=BLADE_DEFAULTS[inp.blade_type]
    phi=defaults['phi']; psi=defaults['psi']; eta=defaults['eta']
    target_v=max(6,min(18,inp.target_flange_velocity_ms))
    vel_p=0.5*rho*target_v**2
    total_p=inp.static_pressure_pa+vel_p
    omega=2*math.pi*inp.rpm/60
    u2=math.sqrt(total_p/max(psi*rho,1e-9))
    d2=2*u2/max(omega,1e-9)
    d1=inp.inlet_diameter_ratio*d2
    b2=inp.outlet_width_ratio*d2
    annulus_area=math.pi*d2*b2
    cm2=q/max(annulus_area,1e-9)
    warnings=[]
    # Keep internal impeller meridional velocity in practical range by adjusting b2, but later score if ratio poor.
    cm2_target=phi*u2
    if cm2>1.25*cm2_target:
        b2=q/(math.pi*d2*cm2_target)
        warnings.append('b₂ was auto-increased to keep internal impeller meridional velocity in practical range.')
        cm2=cm2_target
    elif cm2<0.55*cm2_target:
        warnings.append('b₂ may be wide for this duty; check efficiency and casing width.')
    b1=1.15*b2
    sigma=stodola_slip_factor(inp.blade_count,inp.beta2_deg,d1/d2)
    air_kw=q*total_p/1000
    shaft_kw=air_kw/max(eta*inp.drive_eff,1e-6)
    motor_required=shaft_kw/max(inp.motor_eff,1e-6)*(1+inp.design_margin/100)
    motor=selected_motor(motor_required)
    h=total_p/(rho*9.80665); n_rps=inp.rpm/60
    ns=n_rps*math.sqrt(q)/max(h**0.75,1e-9)
    ds=d2*max(h**0.25,1e-9)/max(math.sqrt(q),1e-9)
    flange_area=q/target_v
    vol_w=max(1.25*b2, 0.20*d2)
    vol_h=flange_area/max(vol_w,1e-9)
    cutoff=0.06*d2
    torque=shaft_kw*1000/max(omega,1e-9)
    shaft_d=((16*torque)/(math.pi*inp.shaft_allow_shear_mpa*1e6))**(1/3)*1000*1.35
    shaft_d=max(20,math.ceil(shaft_d/5)*5)
    mat=MATERIALS[inp.material]
    if u2>mat['max_tip']: warnings.append(f'Tip speed {u2:.1f} m/s exceeds preliminary material limit {mat["max_tip"]} m/s.')
    if inp.blade_type=='Forward Curved' and inp.static_pressure_pa>1000: warnings.append('Forward-curved wheel is not preferred for this pressure; backward-curved is usually safer.')
    d1d2=d1/d2; b2d2=b2/d2
    if b2d2>0.28: warnings.append('b₂/D₂ is high. Consider lower RPM/larger wheel, DIDW wheel or multiple fans in parallel.')
    if b2d2<0.05: warnings.append('b₂/D₂ is very low. Passage may be narrow and losses/noise may increase.')
    if target_v>16: warnings.append('High outlet flange velocity may increase duct loss and noise.')
    if target_v<7: warnings.append('Low outlet velocity gives large outlet flange and casing.')
    t=inp.blade_thickness_mm/1000
    disc_area=math.pi*(d2**2-(0.35*d1)**2)/4
    blade_area=(d2-d1)*b2*1.25
    mass=mat['density']*(2*disc_area*max(t,0.003)+inp.blade_count*blade_area*t)*1.25
    # temporary for scoring/sound
    dummy=type('R',(),{})()
    for k,v in dict(shaft_power_kw=shaft_kw,tip_speed_ms=u2,total_pressure_pa=total_p,flange_outlet_velocity_ms=target_v,impeller_od_mm=d2*1000,impeller_id_mm=d1*1000,blade_count=inp.blade_count).items(): setattr(dummy,k,v)
    pc=blade_pitch_chord_ratio(dummy)
    snd=estimate_sound_db(inp, None, shaft_kw, u2, total_p, target_v, pc)
    risk, risk_notes, vib_score=vibration_risk_raw(inp,u2,target_v,pc,cutoff/d2)
    practical_penalty=practicality_penalty(inp,d1d2,b2d2,pc,u2,target_v)
    score=shaft_kw*1.0 + max(0,snd-75)*1.5 + vib_score*8 + practical_penalty*15
    return DesignResult(q,rho,inp.static_pressure_pa,vel_p,total_p,inp.rpm,u2,d2*1000,d1*1000,b2*1000,b1*1000,inp.beta1_deg,inp.beta2_deg,inp.blade_count,sigma,phi,psi,eta,air_kw,shaft_kw,motor,ns,ds,cm2,target_v,vol_w*1000,vol_h*1000,cutoff*1000,shaft_d,mass,snd,risk,vib_score,score,warnings)

def blade_pitch_chord_ratio(res)->float:
    r_mid=0.25*(res.impeller_od_mm+res.impeller_id_mm)
    pitch=math.pi*(2*r_mid)/max(res.blade_count,1)
    chord=0.55*(res.impeller_od_mm-res.impeller_id_mm)
    return pitch/max(chord,1e-6)

def vibration_risk_raw(inp:DutyInput, tip:float, vout:float, pc:float, cutoff_ratio:float)->Tuple[str,List[str],float]:
    score=0; notes=[]
    if tip>70: score+=2; notes.append('High tip speed: balance sensitivity is high.')
    elif tip>55: score+=1; notes.append('Moderate-high tip speed: dynamic balancing is important.')
    if inp.blade_type=='Forward Curved': score+=1; notes.append('Many narrow passages: dust deposits can create unbalance.')
    if inp.blade_type=='Radial Blade': score+=1.5; notes.append('Radial discharge is more pulsating/noisy.')
    if pc<0.65: score+=1; notes.append('Blade pitch is tight; blade-passing noise risk increases.')
    if pc>1.5: score+=0.7; notes.append('Blade pitch is wide; flow guidance/slip may worsen.')
    if vout>16: score+=1; notes.append('High outlet velocity can excite duct turbulence.')
    if cutoff_ratio<0.05: score+=0.7; notes.append('Small cutoff clearance can raise blade-passing pulsation.')
    if not notes: notes.append('No major preliminary vibration flags. Balance to AMCA/ISO practice.')
    return ('Low' if score<=1 else 'Medium' if score<=3 else 'High'), notes, score

def vibration_risk(inp:DutyInput,res:DesignResult)->Tuple[str,List[str]]:
    r,n,s=vibration_risk_raw(inp,res.tip_speed_ms,res.flange_outlet_velocity_ms,blade_pitch_chord_ratio(res),res.volute_cutoff_clearance_mm/res.impeller_od_mm)
    return r,n

def practicality_penalty(inp:DutyInput,d1d2:float,b2d2:float,pc:float,tip:float,vout:float)->float:
    p=0
    if not 0.45<=d1d2<=0.70: p+=abs(d1d2-0.575)*4
    if not 0.06<=b2d2<=0.25: p+=abs(b2d2-0.15)*5
    if not 0.70<=pc<=1.30: p+=abs(pc-1.0)*1.5
    if tip>85: p+=(tip-85)/10
    if not 8<=vout<=14: p+=abs(vout-11)/10
    if inp.blade_type=='Forward Curved' and inp.static_pressure_pa>1000: p+=3
    return p

def make_trial_input(base:DutyInput, blade_type:str, beta2:float, z:int, d1d2:float, b2d2:float, vout:float)->DutyInput:
    return replace(base, blade_type=blade_type, beta1_deg=BLADE_DEFAULTS[blade_type]['beta1'], beta2_deg=beta2, blade_count=z, inlet_diameter_ratio=d1d2, outlet_width_ratio=b2d2, target_flange_velocity_ms=vout)

def optimise_geometry(base:DutyInput)->Tuple[DutyInput,DesignResult,pd.DataFrame]:
    rows=[]; best=None
    candidate_types=list(BLADE_DEFAULTS.keys()) if base.blade_type=='Auto select best' else [base.blade_type]
    for bt in candidate_types:
        defs=BLADE_DEFAULTS[bt]
        for beta2 in defs['beta2_range']:
            for z in defs['z_range']:
                for d1d2 in [0.45,0.50,0.55,0.60,0.65]:
                    for b2d2 in [0.08,0.10,0.12,0.15,0.18,0.22,0.25]:
                        for vout in [8,10,12,14]:
                            trial=make_trial_input(base,bt,beta2,z,d1d2,b2d2,vout)
                            res=design_blower(trial)
                            pc=blade_pitch_chord_ratio(res); b2r=res.outlet_width_mm/res.impeller_od_mm; d1r=res.impeller_id_mm/res.impeller_od_mm
                            feasible=(0.05<=b2r<=0.30 and 0.42<=d1r<=0.72 and 0.55<=pc<=1.60 and res.tip_speed_ms<=MATERIALS[trial.material]['max_tip'])
                            score=res.design_score+(0 if feasible else 100)
                            rows.append({"Blade type":bt,"β1":trial.beta1_deg,"β2":beta2,"Blades":z,"D1/D2":d1r,"b2/D2":b2r,"Pitch/chord":pc,"Outlet velocity m/s":res.flange_outlet_velocity_ms,"Impeller OD mm":res.impeller_od_mm,"b2 mm":res.outlet_width_mm,"Shaft kW":res.shaft_power_kw,"Sound dB(A)":res.sound_db_a_1m,"Vibration":res.vibration_risk,"Feasible":feasible,"Score":score})
                            if best is None or score<best[0]: best=(score,trial,res)
    df=pd.DataFrame(rows).sort_values(['Feasible','Score'],ascending=[False,True]).head(50)
    return best[1],best[2],df

# ---------- geometry and images ----------
def blade_centerline_points(res:DesignResult,n:int=70)->List[Tuple[float,float]]:
    r1=res.impeller_id_mm/2; r2=res.impeller_od_mm/2
    beta1=math.radians(max(5,min(175,res.beta1_deg))); beta2=math.radians(max(5,min(175,res.beta2_deg)))
    rs=np.linspace(r1,r2,n); th=np.zeros_like(rs)
    for i in range(1,n):
        rmid=(rs[i]+rs[i-1])/2; frac=(rmid-r1)/max(r2-r1,1e-9); beta=beta1+(beta2-beta1)*frac
        th[i]=th[i-1]+(1/math.tan(beta))*(rs[i]-rs[i-1])/max(rmid,1e-9)
    return [(float(r*math.cos(t)),float(r*math.sin(t))) for r,t in zip(rs,th)]

def volute_spiral_points(res:DesignResult,n:int=160)->List[Tuple[float,float]]:
    r_imp=res.impeller_od_mm/2; c=res.volute_cutoff_clearance_mm; w=max(res.volute_outlet_width_mm,1); area=res.volute_outlet_width_mm*res.volute_outlet_height_mm
    pts=[]
    for th in np.linspace(math.radians(10),math.radians(360),n):
        r=r_imp+c+(area*(th/(2*math.pi)))/w
        pts.append((float(r*math.cos(th)),float(r*math.sin(th))))
    return pts

def impeller_figure(res:DesignResult):
    blade=blade_centerline_points(res); vol=np.array(volute_spiral_points(res))
    fig,ax=plt.subplots(figsize=(7,7)); r2=res.impeller_od_mm/2; r1=res.impeller_id_mm/2
    ax.add_patch(plt.Circle((0,0),r2,fill=False,lw=2)); ax.add_patch(plt.Circle((0,0),r1,fill=False,lw=1.5)); ax.add_patch(plt.Circle((0,0),max(res.shaft_diameter_mm*1.6,r1*.25),fill=False,ls='--'))
    for k in range(res.blade_count):
        a=2*math.pi*k/res.blade_count; ca,sa=math.cos(a),math.sin(a); pts=np.array([(x*ca-y*sa,x*sa+y*ca) for x,y in blade]); ax.plot(pts[:,0],pts[:,1],lw=.8)
    ax.plot(vol[:,0],vol[:,1],lw=2)
    ax.annotate(f'D₂ {res.impeller_od_mm:.0f} mm',xy=(0,-r2),xytext=(0,-r2*1.25),ha='center',arrowprops=dict(arrowstyle='<->'))
    ax.annotate(f'D₁ {res.impeller_id_mm:.0f} mm',xy=(r1,0),xytext=(r1*1.25,r1*.35),arrowprops=dict(arrowstyle='->'))
    ax.set_aspect('equal'); ax.grid(True); ax.set_xlabel('mm'); ax.set_ylabel('mm'); ax.set_title('Impeller + preliminary volute')
    return fig

def blade_figure(res:DesignResult):
    pts=np.array(blade_centerline_points(res)); fig,ax=plt.subplots(figsize=(7,4)); ax.plot(pts[:,0],pts[:,1],lw=2)
    pc=blade_pitch_chord_ratio(res); rmid=.25*(res.impeller_od_mm+res.impeller_id_mm); pitch=math.pi*(2*rmid)/res.blade_count; chord=.55*(res.impeller_od_mm-res.impeller_id_mm)
    ax.set_aspect('equal'); ax.grid(True); ax.set_title(f'Blade profile: β₁={res.beta1_deg:.1f}°, β₂={res.beta2_deg:.1f}°, pitch={pitch:.0f} mm, chord≈{chord:.0f} mm, pitch/chord={pc:.2f}')
    ax.set_xlabel('mm'); ax.set_ylabel('mm'); return fig

def fig_png_bytes(fig)->bytes:
    bio=io.BytesIO(); fig.savefig(bio,format='png',dpi=160,bbox_inches='tight'); plt.close(fig); return bio.getvalue()

def create_dxf(res:DesignResult)->bytes:
    if not HAS_EZDXF: return b'Install ezdxf to generate DXF files.'
    doc=ezdxf.new('R2010'); msp=doc.modelspace(); r2=res.impeller_od_mm/2; r1=res.impeller_id_mm/2; hub=max(res.shaft_diameter_mm*1.6,r1*.25)
    msp.add_circle((0,0),r2); msp.add_circle((0,0),r1); msp.add_circle((0,0),hub)
    blade=blade_centerline_points(res)
    for k in range(res.blade_count):
        a=2*math.pi*k/res.blade_count; ca,sa=math.cos(a),math.sin(a); msp.add_lwpolyline([(x*ca-y*sa,x*sa+y*ca) for x,y in blade])
    msp.add_lwpolyline(volute_spiral_points(res))
    x0=r2+res.volute_cutoff_clearance_mm; y0=r2*.15; W=res.volute_outlet_width_mm; H=res.volute_outlet_height_mm
    msp.add_lwpolyline([(x0,y0),(x0+H,y0),(x0+H,y0+W),(x0,y0+W),(x0,y0)])
    s=io.StringIO(); doc.write(s); return s.getvalue().encode()

# ---------- tables and reports ----------
def practicality_table(inp:DutyInput,res:DesignResult)->pd.DataFrame:
    pc=blade_pitch_chord_ratio(res); d1d2=res.impeller_id_mm/res.impeller_od_mm; b2d2=res.outlet_width_mm/res.impeller_od_mm
    def stat(ok): return 'OK' if ok else 'REVIEW'
    return pd.DataFrame([
        ['β₂ outlet blade angle',f'{res.beta2_deg:.1f}°','BC 25-50°, FC 110-150°, Radial ≈90°',stat((inp.blade_type.startswith('Backward') and 25<=res.beta2_deg<=50) or (inp.blade_type.startswith('Forward') and 110<=res.beta2_deg<=150) or (inp.blade_type.startswith('Radial') and 85<=res.beta2_deg<=95)),'Controls pressure, power curve and stability.'],
        ['β₁ inlet blade angle',f'{res.beta1_deg:.1f}°','20-45° preliminary',stat(20<=res.beta1_deg<=45),'Controls inlet shock loss and noise.'],
        ['D₁/D₂ inlet ratio',f'{d1d2:.2f}','0.45-0.70',stat(0.45<=d1d2<=0.70),'Too high reduces pressure; too low restricts inlet.'],
        ['b₂/D₂ width ratio',f'{b2d2:.2f}','0.06-0.25 normal; >0.30 usually impractical',stat(0.06<=b2d2<=0.25),'High value suggests DIDW, lower RPM/larger wheel or parallel fans.'],
        ['Blade pitch/chord',f'{pc:.2f}','0.70-1.30 preferred',stat(0.70<=pc<=1.30),'Too close raises blockage/noise; too wide gives poor guidance/slip.'],
        ['Flange outlet velocity',f'{res.flange_outlet_velocity_ms:.1f} m/s','8-14 m/s preferred; <16 acceptable',stat(8<=res.flange_outlet_velocity_ms<=16),'High velocity adds duct loss and sound.'],
        ['Impeller meridional velocity',f'{res.impeller_meridional_velocity_ms:.1f} m/s','Typically 6-14 m/s',stat(6<=res.impeller_meridional_velocity_ms<=16),'Internal passage velocity through wheel.'],
        ['Tip speed',f'{res.tip_speed_ms:.1f} m/s','Below material limit',stat(res.tip_speed_ms<=MATERIALS[inp.material]['max_tip']),'Affects pressure, stress, sound and balancing.'],
    ],columns=['Input / Check','Your value','Practical guide','Status','Meaning'])

def recommendations(inp:DutyInput,res:DesignResult)->pd.DataFrame:
    rec=[]; pc=blade_pitch_chord_ratio(res); b2d2=res.outlet_width_mm/res.impeller_od_mm
    if b2d2>0.25: rec.append(['b₂/D₂ high','Use lower RPM/larger impeller, DIDW construction or multiple fans in parallel.'])
    if res.flange_outlet_velocity_ms>16: rec.append(['Outlet velocity high','Increase outlet flange area or reduce target outlet velocity.'])
    if pc<0.70: rec.append(['Blade pitch too close','Reduce blade count or increase diameter; check blade blockage and blade-passing noise.'])
    if pc>1.30: rec.append(['Blade pitch wide','Increase blade count or improve blade guidance to reduce slip.'])
    if inp.blade_type=='Forward Curved' and inp.static_pressure_pa>1000: rec.append(['Fan type','Backward-curved is usually preferred for this pressure and non-overloading behaviour.'])
    if res.vibration_risk!='Low': rec.append(['Vibration','Increase cutoff clearance, improve inlet/outlet ducting and specify dynamic balancing.'])
    if not rec: rec.append(['Design direction','Preliminary geometry is in practical range. Proceed to detailed CAD/CFD/prototype testing.'])
    return pd.DataFrame(rec,columns=['Issue','Corrective action'])

def performance_curve(res:DesignResult)->pd.DataFrame:
    rows=[]
    for x in np.linspace(.35,1.25,25):
        q=res.q_m3s*x; p=res.total_pressure_pa*max(.05,1.18-.18*x-.18*x*x); eff=res.estimated_total_eff*max(.15,1-1.8*(x-1)**2); power=q*p/1000/max(eff,.05)
        rows.append({'Flow_m3s':q,'Flow_m3h':q*3600,'TotalPressure_Pa':p,'Efficiency':eff,'ShaftPower_kW':power})
    return pd.DataFrame(rows)

def bom_table(inp:DutyInput,res:DesignResult)->pd.DataFrame:
    return pd.DataFrame([
        ['Impeller back plate',inp.material,inp.blade_thickness_mm,'1'],['Impeller shroud/front ring',inp.material,inp.blade_thickness_mm,'1'],['Blades',inp.material,inp.blade_thickness_mm,str(res.blade_count)],['Volute casing',inp.material,inp.casing_thickness_mm,'1 set'],['Shaft','EN8/C45','-',f'Ø{res.shaft_diameter_mm:.0f} mm preliminary'],['Motor','IE3/IE4 TEFC','-',f'{res.selected_motor_kw:.1f} kW']
    ],columns=['Item','Material','Thickness mm','Approx Qty'])

def create_excel(inp,res,opt_df=None)->bytes:
    """Create Excel safely using openpyxl directly.
    This avoids the pandas/openpyxl 'At least one sheet must be visible'
    error seen on Streamlit Cloud / Python 3.14.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    bio=io.BytesIO()
    wb=Workbook()
    # Keep default sheet and rename it, so at least one visible sheet always exists.
    ws=wb.active
    ws.title='Inputs'

    def write_df(ws, df):
        # header
        for c, col in enumerate(df.columns, start=1):
            cell=ws.cell(row=1, column=c, value=str(col))
            cell.font=Font(bold=True)
            cell.fill=PatternFill('solid', fgColor='DDDDDD')
            cell.alignment=Alignment(wrap_text=True, vertical='top')
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, val in enumerate(row, start=1):
                if isinstance(val, (list, tuple, dict)):
                    val=str(val)
                ws.cell(row=r_idx, column=c_idx, value=val)
        for col_cells in ws.columns:
            max_len=10
            col_letter=col_cells[0].column_letter
            for cell in col_cells:
                try:
                    max_len=max(max_len, min(60, len(str(cell.value))))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width=max_len+2

    def add_sheet(name, df):
        # Sheet names must be <=31 chars and unique.
        name=name[:31]
        if name in wb.sheetnames:
            base=name[:28]
            i=1
            while f'{base}_{i}' in wb.sheetnames:
                i+=1
            name=f'{base}_{i}'
        ws=wb.create_sheet(name)
        write_df(ws, df)
        return ws

    inputs_df=pd.DataFrame([asdict(inp)]).T.reset_index().rename(columns={'index':'Input',0:'Value'})
    # Do not show total pressure as a user input; it is calculated.
    inputs_df=inputs_df[inputs_df['Input']!='total_pressure_pa']
    write_df(ws, inputs_df)

    results_dict=asdict(res).copy()
    # Convert warnings list into readable text for Excel.
    if 'warnings' in results_dict:
        results_dict['warnings']='; '.join(results_dict.get('warnings') or [])
    add_sheet('Results', pd.DataFrame(list(results_dict.items()), columns=['Result','Value']))
    add_sheet('Practicality', practicality_table(inp,res))
    add_sheet('Recommendations', recommendations(inp,res))
    add_sheet('Performance Curve', performance_curve(res))
    add_sheet('BOM', bom_table(inp,res))
    if opt_df is not None and hasattr(opt_df, 'empty') and not opt_df.empty:
        add_sheet('Optimisation Options', opt_df)
    elif opt_df is not None:
        add_sheet('Optimisation Options', pd.DataFrame([['No optimisation options generated']], columns=['Message']))
    add_sheet('Manifest', pd.DataFrame([
        ['PDF report'],['Excel calculation'],['DXF drawing'],['Impeller PNG'],['Blade profile PNG'],['Ansys workflow']
    ],columns=['ZIP Manifest']))

    # Ensure at least one visible worksheet.
    for sh in wb.worksheets:
        sh.sheet_state='visible'
    wb.active=0
    wb.save(bio)
    return bio.getvalue()

def create_pdf(inp,res)->bytes:
    if not HAS_REPORTLAB: return b'Install reportlab to generate PDF reports.'
    bio=io.BytesIO(); doc=SimpleDocTemplate(bio,pagesize=A4); styles=getSampleStyleSheet(); story=[Paragraph('Centrifugal Blower Preliminary Design Report v12',styles['Title']),Spacer(1,8),Paragraph('Static pressure input only. Total pressure is calculated from outlet velocity pressure. Preliminary design for engineering review and prototype validation.',styles['BodyText']),Spacer(1,10)]
    main=[['Parameter','Value'],['Blade type',inp.blade_type],['Airflow',f'{inp.airflow_m3h:,.0f} m³/h ({res.q_m3s:.3f} m³/s)'],['Static pressure',f'{res.static_pressure_pa:.0f} Pa'],['Velocity pressure',f'{res.velocity_pressure_pa:.0f} Pa'],['Calculated total pressure',f'{res.total_pressure_pa:.0f} Pa'],['Density',f'{res.density_kgm3:.3f} kg/m³'],['RPM',f'{res.rpm:.0f}'],['Impeller OD D₂',f'{res.impeller_od_mm:.1f} mm'],['Inlet diameter D₁',f'{res.impeller_id_mm:.1f} mm'],['Outlet width b₂',f'{res.outlet_width_mm:.1f} mm'],['β₁ / β₂',f'{res.beta1_deg:.1f}° / {res.beta2_deg:.1f}°'],['Blades',str(res.blade_count)],['Flange outlet W × H',f'{res.volute_outlet_width_mm:.0f} × {res.volute_outlet_height_mm:.0f} mm'],['Flange outlet velocity',f'{res.flange_outlet_velocity_ms:.1f} m/s'],['Sound estimate',f'{res.sound_db_a_1m:.1f} dB(A) at 1 m'],['Vibration risk',res.vibration_risk],['Shaft power',f'{res.shaft_power_kw:.2f} kW'],['Selected motor',f'{res.selected_motor_kw:.1f} kW']]
    tbl=Table(main,colWidths=[180,300]); tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.lightgrey),('GRID',(0,0),(-1,-1),.3,colors.grey),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold')]))
    story += [tbl,Spacer(1,8),Paragraph('Practicality Checks',styles['Heading2'])]
    pt=[list(practicality_table(inp,res).columns)]+practicality_table(inp,res).values.tolist(); t2=Table(pt,colWidths=[110,70,120,55,120]); t2.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.25,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),7)])); story.append(t2)
    story += [Spacer(1,8),Paragraph('Recommendations',styles['Heading2'])]
    for _,r in recommendations(inp,res).iterrows(): story.append(Paragraph(f'• <b>{r["Issue"]}</b>: {r["Corrective action"]}',styles['BodyText']))
    try:
        ib=io.BytesIO(fig_png_bytes(impeller_figure(res))); bb=io.BytesIO(fig_png_bytes(blade_figure(res))); story += [Spacer(1,8),Image(ib,width=260,height=260),Spacer(1,5),Image(bb,width=350,height=190)]
    except Exception: pass
    doc.build(story); return bio.getvalue()

def ansys_workflow_text():
    return 'Export separate solids for Fluent: impeller_solid.step, stationary_volute.step, inlet_duct.step, outlet_duct.step and fluid_domain.step. Use rotating MRF region around impeller and stationary volute region. Mesh blade edges, tongue and inflation layers.'

def make_zip(inp,res,opt_df=None):
    bio=io.BytesIO()
    with zipfile.ZipFile(bio,'w',zipfile.ZIP_DEFLATED) as z:
        z.writestr('blower_design_report.pdf',create_pdf(inp,res)); z.writestr('blower_design_calculations.xlsx',create_excel(inp,res,opt_df)); z.writestr('blower_2d_manufacturing.dxf',create_dxf(res)); z.writestr('impeller_sketch.png',fig_png_bytes(impeller_figure(res))); z.writestr('blade_profile_sketch.png',fig_png_bytes(blade_figure(res))); z.writestr('README_ANSYS_WORKFLOW.txt',ansys_workflow_text()); z.writestr('step_export_note.txt','Real STEP export requires CadQuery/FreeCAD server setup; next module will generate impeller, volute and fluid-domain STEP files.')
    return bio.getvalue()

# ---------- Streamlit UI ----------
st.set_page_config(page_title='Centrifugal Blower Design Toolkit',layout='wide')
def _get_app_password():
    try:
        if 'APP_PASSWORD' in st.secrets: return str(st.secrets['APP_PASSWORD'])
        if 'auth' in st.secrets and 'password' in st.secrets['auth']: return str(st.secrets['auth']['password'])
    except Exception: return ''
    return ''
_pw=_get_app_password()
if _pw:
    st.sidebar.header('Login'); ent=st.sidebar.text_input('Password',type='password')
    if ent!=_pw: st.warning('Enter password in sidebar to continue.'); st.stop()

st.title('Centrifugal Blower Design & Manufacturing Toolkit v12')
st.success('v11: auto-optimised geometry + practical design scoring + report/ZIP outputs aligned with UI')
with st.sidebar:
    st.header('Duty Inputs')
    airflow=st.number_input('Airflow (m³/h)',min_value=100.0,value=40000.0,step=500.0)
    sp=st.number_input('Static pressure (Pa)',min_value=10.0,value=1400.0,step=50.0)
    rpm=st.number_input('Fan speed (RPM)',min_value=100.0,value=800.0,step=50.0)
    geom_mode=st.radio('Geometry mode',['Auto optimise geometry','Manual override'],index=0)
    blade_choice=st.selectbox('Blade type',['Auto select best']+list(BLADE_DEFAULTS.keys()))
    st.header('Air Properties')
    temp_c=st.number_input('Air temperature (°C)',value=35.0,step=1.0); altitude=st.number_input('Altitude (m)',value=0.0,step=100.0)
    auto_density=standard_air_density(temp_c,altitude); use_auto=st.checkbox(f'Use calculated density ({auto_density:.3f} kg/m³)',value=True); density=auto_density if use_auto else st.number_input('Air density (kg/m³)',value=1.20,step=.01)
    st.header('Outlet and Mechanical')
    target_v=st.number_input('Target blower outlet velocity (m/s)',min_value=6.0,max_value=20.0,value=12.0,step=1.0)
    material=st.selectbox('Impeller/casing material',list(MATERIALS.keys()))
    drive_type=st.selectbox('Drive type',['Direct drive','Belt drive','Coupling drive']); drive_eff=st.number_input('Drive efficiency',min_value=.70,max_value=1.0,value=.95 if drive_type=='Belt drive' else .98,step=.01)
    motor_eff=st.number_input('Motor efficiency',min_value=.70,max_value=.99,value=.90,step=.01); margin=st.number_input('Motor design margin (%)',min_value=0.0,value=15.0,step=1.0)
    blade_thk=st.number_input('Blade/disc thickness (mm)',min_value=1.0,value=3.0,step=.5); casing_thk=st.number_input('Casing thickness (mm)',min_value=1.0,value=3.0,step=.5); shaft_tau=st.number_input('Allowable shaft shear stress (MPa)',min_value=20.0,value=40.0,step=5.0)
    if geom_mode=='Manual override':
        bt=blade_choice if blade_choice!='Auto select best' else 'Backward Curved / Backward Inclined'; defs=BLADE_DEFAULTS[bt]
        st.header('Manual Geometry')
        beta2=st.number_input('Outlet blade angle β₂ (deg)',value=float(defs['beta2']),step=1.0); beta1=st.number_input('Inlet blade angle β₁ (deg)',value=float(defs['beta1']),step=1.0); blades=st.number_input('Number of blades',min_value=3,value=int(defs['z']),step=1); b2r=st.number_input('Outlet width ratio b₂/D₂',min_value=.03,max_value=.50,value=.12,step=.01); d1r=st.number_input('Inlet diameter ratio D₁/D₂',min_value=.25,max_value=.85,value=.55,step=.01)
    else:
        bt=blade_choice; beta1=24; beta2=34; blades=12; b2r=.12; d1r=.55

base=DutyInput(airflow,sp,temp_c,altitude,density,rpm,bt,drive_type,motor_eff,drive_eff,margin,beta2,beta1,int(blades),b2r,d1r,target_v,material,blade_thk,casing_thk,shaft_tau)
opt_df=None
if geom_mode=='Auto optimise geometry':
    inp,res,opt_df=optimise_geometry(base)
else:
    inp=base; res=design_blower(inp)

c1,c2,c3,c4,c5=st.columns(5); c1.metric('Impeller OD',f'{res.impeller_od_mm:.0f} mm'); c2.metric('b₂/D₂',f'{res.outlet_width_mm/res.impeller_od_mm:.2f}'); c3.metric('Outlet velocity',f'{res.flange_outlet_velocity_ms:.1f} m/s'); c4.metric('Shaft power',f'{res.shaft_power_kw:.1f} kW'); c5.metric('Sound',f'{res.sound_db_a_1m:.1f} dB(A)')
st.caption(f'Static pressure {res.static_pressure_pa:.0f} Pa + velocity pressure {res.velocity_pressure_pa:.0f} Pa = calculated total pressure {res.total_pressure_pa:.0f} Pa. Selected blade: {inp.blade_type}.')
for w in res.warnings: st.warning(w)

tabs=st.tabs(['Corrective Dashboard','Design Summary','Optimisation Options','Practicality','Sound/Vibration','Curves','Geometry','Exports'])
with tabs[0]:
    st.subheader('Engineering corrective action dashboard')
    st.dataframe(recommendations(inp,res),use_container_width=True)
    bad=practicality_table(inp,res).query("Status == 'REVIEW'")
    if len(bad): st.error('Some design checks need review before manufacturing.'); st.dataframe(bad,use_container_width=True)
    else: st.success('All preliminary practicality checks are within selected guide ranges.')
with tabs[1]:
    st.dataframe(pd.DataFrame([asdict(res)]).T.reset_index().rename(columns={'index':'Result',0:'Value'}),use_container_width=True)
with tabs[2]:
    if opt_df is not None: st.dataframe(opt_df,use_container_width=True)
    else: st.info('Switch to Auto optimise geometry to compare alternatives.')
with tabs[3]: st.dataframe(practicality_table(inp,res),use_container_width=True)
with tabs[4]:
    st.metric('Estimated sound at 1 m',f'{res.sound_db_a_1m:.1f} dB(A)'); st.metric('Vibration risk',res.vibration_risk)
    for n in vibration_risk(inp,res)[1]: st.write('- '+n)
with tabs[5]:
    curve=performance_curve(res); fig,ax=plt.subplots(); ax.plot(curve.Flow_m3h,curve.TotalPressure_Pa); ax.set_xlabel('Flow m³/h'); ax.set_ylabel('Total Pressure Pa'); ax.grid(True); st.pyplot(fig)
    fig2,ax2=plt.subplots(); ax2.plot(curve.Flow_m3h,curve.ShaftPower_kW); ax2.set_xlabel('Flow m³/h'); ax2.set_ylabel('Shaft Power kW'); ax2.grid(True); st.pyplot(fig2); st.dataframe(curve,use_container_width=True)
with tabs[6]:
    st.pyplot(impeller_figure(res)); st.pyplot(blade_figure(res)); st.info(f'Outlet flange: {res.volute_outlet_width_mm:.0f} mm × {res.volute_outlet_height_mm:.0f} mm; outlet velocity {res.flange_outlet_velocity_ms:.1f} m/s')
with tabs[7]:
    st.download_button('Download PDF Report',create_pdf(inp,res),'blower_design_report.pdf')
    st.download_button('Download Excel Calculation',create_excel(inp,res,opt_df),'blower_design_calculations.xlsx')
    st.download_button('Download DXF Drawing',create_dxf(res),'blower_2d_manufacturing.dxf')
    st.download_button('Download Complete ZIP Package',make_zip(inp,res,opt_df),'blower_design_package.zip')
